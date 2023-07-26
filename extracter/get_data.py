import openpyxl
import psycopg2
from utils.check import get_date, three_months, which_metric, get_multiplier
from pprint import pprint
from utils.check_document_name import year_from_name


def from_excel(sheet, file_name):

    print(type(sheet))
    
    position_three_months = None
    position_six_months = None

    income_statement = {}

    year = year_from_name(file_name)

    mulitplier = get_multiplier(sheet)

    for row in sheet.iter_rows(values_only=True):

        for i, value in enumerate(row): 
            if three_months(str(value)) == "three":
                position_three_months = i
            if three_months(str(value)) == "six":
                position_six_months = i

    # if position_six_months < position_three_months:
    #     print("test")

        #fix this to make sure the correct value is always reterieved. 

    for row in sheet.iter_rows(values_only=True):
        for i, value in enumerate(row): 
            if isinstance(value, str):
                if get_date(value) != False:
                    if get_date(value).year == 2023:
                        print(get_date(value), i, position_three_months)   

    if position_three_months:
        for row in sheet.iter_rows(values_only=True):
            metric = ""
            for i in row: 
                if isinstance(i, str):
                    metric = which_metric(i) 
                    if isinstance(row[position_three_months], int):
                        income_statement[metric] = row[position_three_months] * mulitplier
                        


    pprint(income_statement)



        
