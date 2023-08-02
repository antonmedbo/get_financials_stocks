import openpyxl
import psycopg2
from utils.check import get_date, three_months, which_metric, get_multiplier
from pprint import pprint
from utils.check_document_name import year_from_name
from dateutil.relativedelta import relativedelta


def from_excel(sheet, year):

    print("test")
    
    # position_three_months = None
    # position_six_months = None

    # income_statement = {}
    # end_date = ""
    # income_statement["start_date"] = "" 
    # income_statement["end_date"] = ""

    # year = int(year)

    # multiplier = get_multiplier(sheet)

    # for row in sheet.iter_rows(values_only=True):

    #     for i, value in enumerate(row): 
    #         if three_months(str(value)) == "three":
    #             position_three_months = i
    #         if three_months(str(value)) == "six":
    #             position_six_months = i

    # # if position_six_months < position_three_months:
    # #     print("test")

    #     #fix this to make sure the correct value is always reterieved. 

    # for row in sheet.iter_rows(values_only=True):
    #     for i, value in enumerate(row):
    #         if isinstance(value, str):
    #             if get_date(value) != False:
    #                 if get_date(value).year == year:
    #                     end_date = get_date(value) 

    # if position_three_months and end_date:
    #     income_statement["start_date"] = end_date - relativedelta(months=3) + relativedelta(days=1)
    #     income_statement["end_date"] = end_date
    #     for row in sheet.iter_rows(values_only=True):
    #         metric = ""
    #         for i in row: 
    #             if isinstance(i, str):
    #                 metric = which_metric(i) 
    #                 if isinstance(row[position_three_months], int):
    #                     income_statement[metric] = row[position_three_months] * multiplier
    
    # return(income_statement)



        
